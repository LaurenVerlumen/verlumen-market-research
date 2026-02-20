"""Import products and categories from the Verlumen Excel spreadsheet."""
from __future__ import annotations

import io
from pathlib import Path
from typing import Union

import openpyxl

from src.services.alibaba_parser import parse_alibaba_url


def parse_excel(source: Union[str, Path, bytes]) -> list[dict]:
    """Parse the Verlumen Product Research Excel file.

    The expected format (sheet "Verlumen Product Research"):
    - Column A: Category name (appears at the start of each group)
    - Column B: Alibaba URL
    - Blank rows separate categories
    - When A has text AND B has URL -> new category + first product
    - When A is empty AND B has URL -> product in current category

    Args:
        source: File path (str/Path) or raw file bytes (e.g. from an upload).

    Returns:
        List of dicts, each with:
            {"category": "name", "products": [{"url": "...", "name": "...", "product_id": "..."}]}
    """
    if isinstance(source, bytes):
        wb = openpyxl.load_workbook(io.BytesIO(source), data_only=True)
    else:
        wb = openpyxl.load_workbook(str(source), data_only=True)

    # Try the expected sheet name, fall back to first sheet
    sheet_name = "Verlumen Product Research"
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]

    categories: list[dict] = []
    current_category: dict | None = None

    for row in ws.iter_rows(min_row=1, max_col=3, values_only=True):
        col_a = row[0] if len(row) > 0 else None
        col_b = row[1] if len(row) > 1 else None
        col_c = row[2] if len(row) > 2 else None

        col_a_str = str(col_a).strip() if col_a else ""
        col_b_str = str(col_b).strip() if col_b else ""
        col_c_str = str(col_c).strip() if col_c else ""

        is_url = col_b_str.startswith("http")

        if col_a_str and is_url:
            # New category with its first product
            current_category = {"category": col_a_str, "products": []}
            categories.append(current_category)
            product = _make_product(col_b_str)
            if col_c_str:
                product["supplier"] = col_c_str
            current_category["products"].append(product)

        elif not col_a_str and is_url:
            # Product in current category
            if current_category is None:
                # Edge case: URL before any category header -- create unnamed group
                current_category = {"category": "Uncategorized", "products": []}
                categories.append(current_category)
            product = _make_product(col_b_str)
            if col_c_str:
                product["supplier"] = col_c_str
            current_category["products"].append(product)

        # Blank rows (both A and B empty) are category separators -- nothing to do

    wb.close()
    return categories


def _make_product(url: str) -> dict:
    """Build a product dict from an Alibaba URL using the parser."""
    parsed = parse_alibaba_url(url)
    return {
        "url": parsed["clean_url"],
        "name": parsed["name"],
        "product_id": parsed["product_id"],
    }
