"""Export research data to a multi-sheet Excel workbook."""
import logging
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Style constants
_HEADER_FONT = Font(bold=True, size=11)
_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_GREEN_FONT = Font(color="006100")
_YELLOW_FONT = Font(color="9C5700")
_RED_FONT = Font(color="9C0006")
_GRAY_FONT = Font(color="999999", italic=True)

# Executive summary styles (Verlumen brand)
_TITLE_FONT = Font(bold=True, size=16, color="1F4E79")
_SUBTITLE_FONT = Font(bold=False, size=11, color="4472C4")
_KPI_HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
_KPI_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_KPI_BORDER = Border(
    left=Side(style="thin", color="1F4E79"),
    right=Side(style="thin", color="1F4E79"),
    top=Side(style="thin", color="1F4E79"),
    bottom=Side(style="thin", color="1F4E79"),
)
_VERLUMEN_BLUE = "1F4E79"
_VERLUMEN_ACCENT = "4472C4"


class ExcelExporter:
    """Export enriched research data to a formatted .xlsx file."""

    def export(
        self,
        products_data: list[dict],
        output_path: str,
        include_ml: bool = True,
        include_profit: bool = True,
    ) -> str:
        """Export research data to an Excel workbook.

        Parameters
        ----------
        products_data : list[dict]
            Each dict should contain: category, name, alibaba_url,
            alibaba_price_min, alibaba_price_max, analysis (dict),
            competitors (list[dict]).  Optionally: ml_data (dict),
            profit_data (dict).
        output_path : str
            Destination file path for the .xlsx file.
        include_ml : bool
            Whether to include ML analysis columns and AI Recommendations sheet.
        include_profit : bool
            Whether to include Profit Analysis sheet.

        Returns
        -------
        str  The absolute path of the created file.
        """
        wb = Workbook()

        # Sheet 1 -- Executive Summary (new: charts + KPIs)
        ws_exec = wb.active
        ws_exec.title = "Executive Summary"
        self._build_executive_summary_sheet(ws_exec, products_data)

        # Sheet 2 -- Summary
        ws_summary = wb.create_sheet("Summary")
        self._build_summary_sheet(ws_summary, products_data, include_ml=include_ml)

        # Sheet 3 -- Detailed Competitors
        ws_detail = wb.create_sheet("Detailed Competitors")
        self._build_detail_sheet(ws_detail, products_data)

        # Sheet 4 -- Category Analysis
        ws_cat = wb.create_sheet("Category Analysis")
        self._build_category_sheet(ws_cat, products_data)

        # Sheet 5 -- Profit Analysis (optional)
        if include_profit:
            ws_profit = wb.create_sheet("Profit Analysis")
            self._build_profit_sheet(ws_profit, products_data)

        # Sheet 6 -- AI Recommendations (optional)
        if include_ml:
            ws_ai = wb.create_sheet("AI Recommendations")
            self._build_ai_recommendations_sheet(ws_ai, products_data)

        # Save
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        logger.info("Excel exported to %s", output_path)
        return str(Path(output_path).resolve())

    # ------------------------------------------------------------------
    # Sheet builders
    # ------------------------------------------------------------------

    def _build_summary_sheet(
        self, ws: Any, products: list[dict], include_ml: bool = True,
    ) -> None:
        headers = [
            "Category",
            "Product Name",
            "Alibaba URL",
            "Alibaba Price Range",
            "Amazon Avg Price",
            "Amazon Median Price",
            "# Competitors",
            "Avg Rating",
            "Avg Reviews",
            "Competition Score",
            "Opportunity Score",
            "Suggested Price Range",
        ]
        if include_ml:
            headers += [
                "Match Score (Best)",
                "Price Strategy",
                "Demand Level",
                "Est Monthly Revenue",
                "Profit Margin %",
            ]
        self._write_header_row(ws, headers)

        for row_idx, p in enumerate(products, start=2):
            a: dict = p.get("analysis") or {}
            is_researched = bool(a and a.get("total_competitors", 0) > 0)

            alibaba_price = self._format_alibaba_price(
                p.get("alibaba_price_min"), p.get("alibaba_price_max")
            )
            suggested = self._format_price_range(
                a.get("suggested_price_min"), a.get("suggested_price_max")
            )

            ws.cell(row=row_idx, column=1, value=p.get("category", ""))
            ws.cell(row=row_idx, column=2, value=p.get("name", ""))
            ws.cell(row=row_idx, column=3, value=p.get("alibaba_url", ""))
            ws.cell(row=row_idx, column=4, value=alibaba_price)

            if is_researched:
                ws.cell(row=row_idx, column=5, value=a.get("price_mean", 0)).number_format = (
                    numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                )
                ws.cell(row=row_idx, column=6, value=a.get("price_median", 0)).number_format = (
                    numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                )
                ws.cell(row=row_idx, column=7, value=a.get("total_competitors", 0))
                ws.cell(row=row_idx, column=8, value=a.get("avg_rating", 0)).number_format = "0.0"
                ws.cell(row=row_idx, column=9, value=a.get("avg_reviews", 0))
                ws.cell(row=row_idx, column=10, value=a.get("competition_score", 0)).number_format = "0.0"
                ws.cell(row=row_idx, column=11, value=a.get("opportunity_score", 0)).number_format = "0.0"
                ws.cell(row=row_idx, column=12, value=suggested)
            else:
                for col in range(5, 13):
                    cell = ws.cell(row=row_idx, column=col, value="Not researched")
                    cell.font = _GRAY_FONT

            # ML columns
            if include_ml:
                ml = p.get("ml_data") or {}
                if is_researched and ml:
                    best_match = ml.get("best_match_score")
                    ws.cell(
                        row=row_idx, column=13,
                        value=best_match if best_match is not None else "",
                    ).number_format = "0.0"
                    ws.cell(row=row_idx, column=14, value=ml.get("price_strategy", ""))
                    ws.cell(row=row_idx, column=15, value=ml.get("demand_level", ""))
                    est_rev = ml.get("estimated_monthly_revenue")
                    ws.cell(
                        row=row_idx, column=16,
                        value=est_rev if est_rev is not None else "",
                    ).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                    margin = ml.get("profit_margin_pct")
                    ws.cell(
                        row=row_idx, column=17,
                        value=margin if margin is not None else "",
                    ).number_format = "0.0"
                else:
                    for col in range(13, 18):
                        cell = ws.cell(
                            row=row_idx, column=col,
                            value="" if is_researched else "Not researched",
                        )
                        if not is_researched:
                            cell.font = _GRAY_FONT

        # Conditional formatting on Opportunity Score (col 11)
        last_row = len(products) + 1
        if last_row >= 2:
            opp_range = f"K2:K{last_row}"
            self._add_traffic_light_formatting(ws, opp_range, thresholds=(60, 30))

            # Conditional formatting on Profit Margin % (col 17) if ML included
            if include_ml:
                margin_range = f"Q2:Q{last_row}"
                self._add_margin_formatting(ws, margin_range)

        self._auto_column_width(ws)

    def _build_executive_summary_sheet(self, ws: Any, products: list[dict]) -> None:
        """Sheet 1: Executive Summary with KPI table and embedded charts."""
        # -- Title row --
        ws.merge_cells("A1:F1")
        title_cell = ws.cell(row=1, column=1, value="Verlumen Market Research Report")
        title_cell.font = _TITLE_FONT
        title_cell.alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:F2")
        date_cell = ws.cell(row=2, column=1, value=f"Generated: {date.today().strftime('%B %d, %Y')}")
        date_cell.font = _SUBTITLE_FONT
        date_cell.alignment = Alignment(horizontal="center")

        # -- KPI calculations --
        researched = [
            p for p in products
            if p.get("analysis") and p["analysis"].get("total_competitors", 0) > 0
        ]
        total_products = len(products)
        avg_opp = 0.0
        avg_comp = 0.0
        total_competitors = 0
        if researched:
            opp_scores = [p["analysis"].get("opportunity_score", 0) for p in researched]
            comp_scores = [p["analysis"].get("competition_score", 0) for p in researched]
            avg_opp = sum(opp_scores) / len(opp_scores)
            avg_comp = sum(comp_scores) / len(comp_scores)
            total_competitors = sum(
                p["analysis"].get("total_competitors", 0) for p in researched
            )

        # -- KPI table at row 4 --
        kpi_headers = ["Metric", "Value"]
        for col_idx, h in enumerate(kpi_headers, start=1):
            cell = ws.cell(row=4, column=col_idx, value=h)
            cell.font = _KPI_HEADER_FONT
            cell.fill = _KPI_HEADER_FILL
            cell.border = _KPI_BORDER
            cell.alignment = Alignment(horizontal="center")

        kpis = [
            ("Total Products", total_products),
            ("Avg Opportunity Score", round(avg_opp, 1)),
            ("Avg Competition Score", round(avg_comp, 1)),
            ("Total Competitors Tracked", total_competitors),
        ]
        for i, (label, value) in enumerate(kpis):
            row = 5 + i
            lbl_cell = ws.cell(row=row, column=1, value=label)
            lbl_cell.font = Font(bold=True, size=11)
            lbl_cell.border = _KPI_BORDER
            val_cell = ws.cell(row=row, column=2, value=value)
            val_cell.border = _KPI_BORDER
            val_cell.alignment = Alignment(horizontal="center")
            if isinstance(value, float):
                val_cell.number_format = "0.0"

        # -- Chart 1: Price Distribution (bar chart) --
        # Build price buckets from competitor prices
        buckets = {"$0-10": 0, "$10-25": 0, "$25-50": 0, "$50-100": 0, "$100+": 0}
        for p in products:
            for c in p.get("competitors", []):
                price = c.get("price")
                if price is None:
                    continue
                if price < 10:
                    buckets["$0-10"] += 1
                elif price < 25:
                    buckets["$10-25"] += 1
                elif price < 50:
                    buckets["$25-50"] += 1
                elif price < 100:
                    buckets["$50-100"] += 1
                else:
                    buckets["$100+"] += 1

        # Write price distribution data to a helper area (row 11+)
        ws.cell(row=11, column=1, value="Price Range").font = Font(bold=True, size=9, color="999999")
        ws.cell(row=11, column=2, value="Count").font = Font(bold=True, size=9, color="999999")
        for i, (bucket, count) in enumerate(buckets.items()):
            ws.cell(row=12 + i, column=1, value=bucket)
            ws.cell(row=12 + i, column=2, value=count)

        chart1 = BarChart()
        chart1.type = "col"
        chart1.title = "Price Distribution"
        chart1.y_axis.title = "Number of Products"
        chart1.x_axis.title = "Price Range"
        chart1.style = 10
        data1 = Reference(ws, min_col=2, min_row=11, max_row=16)
        cats1 = Reference(ws, min_col=1, min_row=12, max_row=16)
        chart1.add_data(data1, titles_from_data=True)
        chart1.set_categories(cats1)
        chart1.shape = 4
        chart1.width = 16
        chart1.height = 10
        # Apply Verlumen blue to the bars
        if chart1.series:
            chart1.series[0].graphicalProperties.solidFill = _VERLUMEN_ACCENT
        ws.add_chart(chart1, "D4")

        # -- Chart 2: Top 10 Products by Opportunity Score --
        sorted_products = sorted(
            researched,
            key=lambda x: x.get("analysis", {}).get("opportunity_score", 0),
            reverse=True,
        )[:10]

        ws.cell(row=18, column=1, value="Product").font = Font(bold=True, size=9, color="999999")
        ws.cell(row=18, column=2, value="Opportunity Score").font = Font(bold=True, size=9, color="999999")
        for i, p in enumerate(sorted_products):
            name = p.get("name", "")
            # Truncate long names for chart readability
            display_name = name[:30] + "..." if len(name) > 30 else name
            ws.cell(row=19 + i, column=1, value=display_name)
            ws.cell(row=19 + i, column=2, value=p["analysis"].get("opportunity_score", 0))

        if sorted_products:
            last_data_row = 18 + len(sorted_products)
            chart2 = BarChart()
            chart2.type = "bar"
            chart2.title = "Top 10 Products by Opportunity Score"
            chart2.y_axis.title = "Product"
            chart2.x_axis.title = "Opportunity Score"
            chart2.style = 10
            data2 = Reference(ws, min_col=2, min_row=18, max_row=last_data_row)
            cats2 = Reference(ws, min_col=1, min_row=19, max_row=last_data_row)
            chart2.add_data(data2, titles_from_data=True)
            chart2.set_categories(cats2)
            chart2.shape = 4
            chart2.width = 16
            chart2.height = 12
            if chart2.series:
                chart2.series[0].graphicalProperties.solidFill = _VERLUMEN_BLUE
            ws.add_chart(chart2, "D18")

        # Set column widths for the KPI area
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 18

    def _build_detail_sheet(self, ws: Any, products: list[dict]) -> None:
        headers = [
            "Our Product",
            "Amazon ASIN",
            "Amazon Title",
            "Price",
            "Rating",
            "Reviews",
            "Bought Last Month",
            "Prime",
            "Badge",
            "Amazon URL",
        ]
        self._write_header_row(ws, headers)

        row = 2
        for p in products:
            name = p.get("name", "")
            competitors = p.get("competitors", [])
            if not competitors:
                # Mark unresearched products
                ws.cell(row=row, column=1, value=name)
                cell = ws.cell(row=row, column=2, value="Not researched")
                cell.font = _GRAY_FONT
                row += 1
                continue
            for c in competitors:
                ws.cell(row=row, column=1, value=name)
                ws.cell(row=row, column=2, value=c.get("asin", ""))
                ws.cell(row=row, column=3, value=c.get("title", ""))
                price = c.get("price")
                cell_price = ws.cell(row=row, column=4, value=price if price is not None else "")
                if price is not None:
                    cell_price.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                ws.cell(row=row, column=5, value=c.get("rating", "")).number_format = "0.0"
                ws.cell(row=row, column=6, value=c.get("review_count", ""))
                ws.cell(row=row, column=7, value=c.get("bought_last_month", ""))
                ws.cell(row=row, column=8, value="Yes" if c.get("is_prime") else "No")
                ws.cell(row=row, column=9, value=c.get("badge", ""))
                ws.cell(row=row, column=10, value=c.get("amazon_url", ""))
                row += 1

        self._auto_column_width(ws)

    def _build_category_sheet(self, ws: Any, products: list[dict]) -> None:
        headers = [
            "Category",
            "Num Products",
            "Avg Competition Score",
            "Avg Opportunity Score",
            "Best Opportunity",
            "Avg Amazon Price",
        ]
        self._write_header_row(ws, headers)

        # Group products by category
        categories: dict[str, list[dict]] = {}
        for p in products:
            cat = p.get("category", "Uncategorized")
            categories.setdefault(cat, []).append(p)

        row = 2
        for cat, items in sorted(categories.items()):
            comp_scores = [
                i.get("analysis", {}).get("competition_score", 0) for i in items
            ]
            opp_scores = [
                i.get("analysis", {}).get("opportunity_score", 0) for i in items
            ]
            avg_prices = [
                i.get("analysis", {}).get("price_mean", 0) for i in items
            ]

            best_item = max(items, key=lambda x: x.get("analysis", {}).get("opportunity_score", 0))

            avg_comp = sum(comp_scores) / len(comp_scores) if comp_scores else 0
            avg_opp = sum(opp_scores) / len(opp_scores) if opp_scores else 0
            avg_price = sum(avg_prices) / len(avg_prices) if avg_prices else 0

            ws.cell(row=row, column=1, value=cat)
            ws.cell(row=row, column=2, value=len(items))
            ws.cell(row=row, column=3, value=round(avg_comp, 1)).number_format = "0.0"
            ws.cell(row=row, column=4, value=round(avg_opp, 1)).number_format = "0.0"
            ws.cell(row=row, column=5, value=best_item.get("name", ""))
            ws.cell(row=row, column=6, value=round(avg_price, 2)).number_format = (
                numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            )
            row += 1

        # Conditional formatting on Avg Opportunity Score (col 4)
        last_row = row - 1
        if last_row >= 2:
            opp_range = f"D2:D{last_row}"
            self._add_traffic_light_formatting(ws, opp_range, thresholds=(60, 30))

        # -- Grouped bar chart: Avg Price + Avg Rating per Category --
        if last_row >= 2:
            # Write helper data for chart (avg price col=8, avg rating col=9)
            # We need the data in adjacent columns for a grouped bar chart
            ws.cell(row=1, column=8, value="Category").font = Font(bold=True, size=9, color="999999")
            ws.cell(row=1, column=9, value="Avg Price ($)").font = Font(bold=True, size=9, color="999999")
            ws.cell(row=1, column=10, value="Avg Rating").font = Font(bold=True, size=9, color="999999")

            chart_row = 2
            for cat, items in sorted(categories.items()):
                avg_prices_list = [
                    i.get("analysis", {}).get("price_mean", 0) for i in items
                ]
                avg_ratings = []
                for i in items:
                    for c in i.get("competitors", []):
                        r = c.get("rating")
                        if r is not None:
                            avg_ratings.append(r)

                avg_price_val = sum(avg_prices_list) / len(avg_prices_list) if avg_prices_list else 0
                avg_rating_val = sum(avg_ratings) / len(avg_ratings) if avg_ratings else 0

                ws.cell(row=chart_row, column=8, value=cat)
                ws.cell(row=chart_row, column=9, value=round(avg_price_val, 2))
                ws.cell(row=chart_row, column=10, value=round(avg_rating_val, 2))
                chart_row += 1

            chart_last_row = chart_row - 1
            if chart_last_row >= 2:
                chart = BarChart()
                chart.type = "col"
                chart.grouping = "clustered"
                chart.title = "Avg Price & Avg Rating by Category"
                chart.style = 10
                chart.y_axis.title = "Value"
                chart.x_axis.title = "Category"

                data_ref = Reference(ws, min_col=9, min_row=1, max_col=10, max_row=chart_last_row)
                cats_ref = Reference(ws, min_col=8, min_row=2, max_row=chart_last_row)
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats_ref)
                chart.shape = 4
                chart.width = 20
                chart.height = 12
                # Color the series
                if len(chart.series) > 0:
                    chart.series[0].graphicalProperties.solidFill = _VERLUMEN_BLUE
                if len(chart.series) > 1:
                    chart.series[1].graphicalProperties.solidFill = _VERLUMEN_ACCENT
                ws.add_chart(chart, "A" + str(last_row + 3))

        self._auto_column_width(ws)

    def _build_profit_sheet(self, ws: Any, products: list[dict]) -> None:
        """Sheet 4: Profit Analysis."""
        headers = [
            "Product",
            "Alibaba Cost",
            "Landed Cost",
            "Amazon Selling Price (Competitive)",
            "Amazon Fee",
            "Net Profit/Unit",
            "Margin %",
            "ROI %",
            "Break-even Units",
            "Monthly Profit Est",
        ]
        self._write_header_row(ws, headers)

        row = 2
        for p in products:
            profit = p.get("profit_data") or {}
            strategies = profit.get("strategies") or {}
            competitive = strategies.get("competitive") or {}
            be = profit.get("break_even_units") or {}
            monthly = profit.get("monthly_profit_estimate") or {}
            comp_monthly = monthly.get("competitive") or {}

            ws.cell(row=row, column=1, value=p.get("name", ""))

            if not strategies:
                # No profit data
                alibaba_price = self._format_alibaba_price(
                    p.get("alibaba_price_min"), p.get("alibaba_price_max")
                )
                ws.cell(row=row, column=2, value=alibaba_price or "N/A")
                cell = ws.cell(row=row, column=3, value="Not available")
                cell.font = _GRAY_FONT
                row += 1
                continue

            ws.cell(
                row=row, column=2,
                value=self._format_alibaba_price(p.get("alibaba_price_min"), p.get("alibaba_price_max")),
            )
            ws.cell(row=row, column=3, value=profit.get("landed_cost", 0)).number_format = (
                numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            )
            ws.cell(row=row, column=4, value=competitive.get("selling_price", 0)).number_format = (
                numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            )
            ws.cell(row=row, column=5, value=competitive.get("amazon_fee", 0)).number_format = (
                numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            )
            ws.cell(row=row, column=6, value=competitive.get("net_profit", 0)).number_format = (
                numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            )
            ws.cell(row=row, column=7, value=competitive.get("profit_margin_pct", 0)).number_format = "0.0"
            ws.cell(row=row, column=8, value=competitive.get("roi_pct", 0)).number_format = "0.0"
            ws.cell(row=row, column=9, value=be.get("competitive", 0))
            ws.cell(row=row, column=10, value=comp_monthly.get("monthly_profit", 0)).number_format = (
                numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            )
            row += 1

        # Conditional formatting on Margin % (col 7)
        last_row = row - 1
        if last_row >= 2:
            margin_range = f"G2:G{last_row}"
            self._add_margin_formatting(ws, margin_range)

        self._auto_column_width(ws)

    def _build_ai_recommendations_sheet(self, ws: Any, products: list[dict]) -> None:
        """Sheet 5: AI Recommendations."""
        headers = [
            "Product",
            "Optimized Search Query",
            "Top 3 Matching Competitors",
            "Recommended Price",
            "Rationale",
            "Market Size",
        ]
        self._write_header_row(ws, headers)

        row = 2
        for p in products:
            ml = p.get("ml_data") or {}
            ws.cell(row=row, column=1, value=p.get("name", ""))

            if not ml:
                cell = ws.cell(row=row, column=2, value="Not researched")
                cell.font = _GRAY_FONT
                row += 1
                continue

            ws.cell(row=row, column=2, value=ml.get("optimized_query", ""))
            ws.cell(row=row, column=3, value=ml.get("top_3_matches", ""))
            rec_price = ml.get("recommended_price")
            ws.cell(
                row=row, column=4,
                value=f"${rec_price:.2f}" if rec_price is not None else "",
            )
            ws.cell(row=row, column=5, value=ml.get("rationale", ""))
            ws.cell(row=row, column=6, value=ml.get("market_size", ""))
            row += 1

        self._auto_column_width(ws)

    # ------------------------------------------------------------------
    # Formatting helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _write_header_row(ws: Any, headers: list[str]) -> None:
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

    @staticmethod
    def _auto_column_width(ws: Any, min_width: int = 10, max_width: int = 50) -> None:
        for col_cells in ws.columns:
            max_len = min_width
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, min(len(val) + 2, max_width))
            ws.column_dimensions[col_letter].width = max_len

    @staticmethod
    def _format_alibaba_price(pmin: Any, pmax: Any) -> str:
        if pmin is not None and pmax is not None:
            return f"${pmin:.2f} - ${pmax:.2f}"
        if pmin is not None:
            return f"${pmin:.2f}"
        if pmax is not None:
            return f"${pmax:.2f}"
        return ""

    @staticmethod
    def _format_price_range(pmin: Any, pmax: Any) -> str:
        if pmin is not None and pmax is not None:
            return f"${pmin:.2f} - ${pmax:.2f}"
        return ""

    @staticmethod
    def _add_traffic_light_formatting(
        ws: Any, cell_range: str, thresholds: tuple[float, float] = (60, 30),
    ) -> None:
        """Add green/yellow/red conditional formatting based on thresholds."""
        high, low = thresholds
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(
                operator="greaterThanOrEqual",
                formula=[str(high)],
                fill=_GREEN_FILL,
                font=_GREEN_FONT,
            ),
        )
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(
                operator="between",
                formula=[str(low), str(high - 0.1)],
                fill=_YELLOW_FILL,
                font=_YELLOW_FONT,
            ),
        )
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(
                operator="lessThan",
                formula=[str(low)],
                fill=_RED_FILL,
                font=_RED_FONT,
            ),
        )

    @staticmethod
    def _add_margin_formatting(ws: Any, cell_range: str) -> None:
        """Add green (>30%), yellow (15-30%), red (<15%) formatting for margin columns."""
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(
                operator="greaterThan",
                formula=["30"],
                fill=_GREEN_FILL,
                font=_GREEN_FONT,
            ),
        )
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(
                operator="between",
                formula=["15", "30"],
                fill=_YELLOW_FILL,
                font=_YELLOW_FONT,
            ),
        )
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(
                operator="lessThan",
                formula=["15"],
                fill=_RED_FILL,
                font=_RED_FONT,
            ),
        )
